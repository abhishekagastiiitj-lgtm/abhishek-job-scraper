"""
Job Scraper for Abhishek Agasti — FULL EDITION v6
==================================================
Profile  : Data Science | Business Analyst | ML/AI | 2-3 yr exp
Target   : Remote / Pan-India roles, associate level (2-3 yrs)
Output   : Daily HTML email  +  jobs_tracker.xlsx (appended each run)
           + Per-job ATS-optimised resume PDFs linked in Excel col N
           (Claude rewrites the base resume for each JD for 90%+ keyword match)

NEW in v6
─────────
• Column N "Tailored Resume" — for every new job, Claude API rewrites
  Abhishek's resume to match the JD keywords exactly (same structure,
  reworded bullets), saves as resume_<job_id>.pdf, and stores a file
  path in the Excel cell.  Download the artifact zip after each run.
• Set ANTHROPIC_API_KEY as a GitHub Secret to enable resume generation.
  Without it the column shows "API key not set" gracefully.
• RESUME_OUTPUT_DIR = "resumes/" — all generated PDFs land here and are
  uploaded as part of the job-tracker-N artifact bundle.

Sources
───────
 ── REMOTE AGGREGATORS ─────────────────────────────────────────────────────
 1. Remotive           — free public API, all companies
 2. Jobicy             — free public API, all companies
 3. We Work Remotely   — RSS feed, curated remote jobs
 4. Remote OK          — free public JSON API, tech remote
 5. Working Nomads     — RSS, remote tech/marketing roles
 6. Jobspresso         — RSS feed, remote tech roles
 7. AngelList/Wellfound— public JSON search, startup roles
 8. Surely             — aggregator: scrapes 500+ job boards simultaneously

 ── ATS BOARDS ─────────────────────────────────────────────────────────────
 9. Greenhouse         — per-board, large seed list (~100 boards)
10. Lever              — per-board, large seed list (~60 boards)
11. Ashby              — GraphQL, per-board seed list
12. Workday            — per-tenant keyword POST search
13. SmartRecruiters    — per-company keyword search

 ── INDIA PORTALS ──────────────────────────────────────────────────────────
14. Instahyre          — public job listings API (India-focused)
15. Naukri             — public job search (India)
16. iimjobs            — public job search (premium India)

Run locally : python scraper.py
Scheduled   : GitHub Actions → .github/workflows/daily_scraper.yml (8 AM IST)
"""

import json, os, re, smtplib, time, csv
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText
from pathlib              import Path
import xml.etree.ElementTree as ET

import requests

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils  import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("  ⚠  openpyxl not installed — Excel output skipped (pip install openpyxl)")

try:
    from reportlab.lib.pagesizes   import letter
    from reportlab.lib.styles      import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units       import inch
    from reportlab.lib             import colors
    from reportlab.platypus        import (SimpleDocTemplate, Paragraph, Spacer,
                                           HRFlowable, Table, TableStyle)
    from reportlab.lib.enums       import TA_LEFT, TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("  ⚠  reportlab not installed — resume PDF generation skipped")

# ═══════════════════════════════════════════════════════════════════════════════
#  PROFILE CONFIG — tuned for Abhishek Agasti's resume
# ═══════════════════════════════════════════════════════════════════════════════

# PRIMARY role keywords — job must match at least one
ROLE_KEYWORDS = [
    "data scientist",       "data science",         "data analyst",
    "analytics engineer",   "analytics manager",    "data engineer",
    "business analyst",     "business intelligence","bi analyst",
    "bi developer",         "product analyst",      "product data analyst",
    "ml engineer",          "machine learning",     "mlops",
    "financial analyst",    "fraud analyst",        "risk analyst",
    "quantitative analyst", "quant analyst",        "fp&a",
    "revenue analyst",      "ai engineer",          "llm engineer",
    "applied scientist",    "decision scientist",   "growth analyst",
    "insights analyst",     "research analyst",     "pricing analyst",
    "credit analyst",       "demand forecast",      "data platform",
]

# TECH keywords — job must match at least one (broad intentionally)
TECH_KEYWORDS = [
    "python",  "sql",       "pyspark",    "spark",       "databricks",
    "azure",   "power bi",  "xgboost",    "forecasting", "arima",
    "prophet", "fraud",     "fintech",    "analytics",   "machine learning",
    "mlflow",  "delta lake","etl",        "pipeline",    "warehouse",
    "langchain","llm",      "rag",        "a/b test",    "tableau",
    "looker",  "dbt",       "snowflake",  "bigquery",    "airflow",
    "kafka",   "redshift",  "pandas",     "numpy",       "scikit",
    "tensorflow","shap",    "deep learning","nlp",       "generative ai",
    "data",    "ml",        "ai",         "model",       "insight",
]

# Job MUST match ≥1 from each group
KEYWORD_GROUPS = [ROLE_KEYWORDS, TECH_KEYWORDS]

# Location filter — at least one must appear in title+location text
LOCATION_KEYWORDS = [
    "india", "remote", "bangalore", "bengaluru", "mumbai", "hyderabad",
    "delhi", "gurgaon", "gurugram", "noida", "pune", "chennai",
    "kolkata", "ahmedabad", "guwahati", "worldwide", "global", "anywhere",
    "asia", "apac", "work from home", "wfh", "pan india", "pan-india",
    "distributed", "fully remote",
]

# EXPERIENCE level filter — skip if NONE of these appear in title/description
# or if a SENIOR-ONLY keyword is found without any associate-level signal
EXPERIENCE_SIGNALS = [
    "associate", "junior", "analyst", "entry", "0-3", "1-3", "2-3",
    "2-4", "1-4", "fresher", "graduate", "ii", "level 2", "l2",
]
SENIOR_ONLY_KEYWORDS = [
    "director", "vp ", "vice president", "head of", "principal",
    "staff ", "distinguished", "fellow", "c-suite", "cto", "cdo",
    "partner", "managing director", "svp", "evp",
]

# ═══════════════════════════════════════════════════════════════════════════════
#  ATS SEED LISTS
# ═══════════════════════════════════════════════════════════════════════════════

GREENHOUSE_SLUGS = [
    "phonepe","groww","paisabazaar","zerodha","cleartax","polygon1",
    "databricks","postman","stripe","chargebee","mixpanel","amplitude",
    "segment","figma","notion","linear","retool","hashicorp","brex","ramp",
    "dbtlabs","fivetran","airbyte","prefect","hightouch","census",
    "rudderstack","metabase","mode","sisense","thoughtspot",
    "shopify","instacart","faire",
    "fractalanalytics","mu-sigma","latentview",
    "cloudflare","mongodb","confluent","cockroachdb","harness",
    "adyen","marqeta","mercury","plaid","kforce",
    "meesho","razorpay","cred","swiggy","zomato","ola","dunzo",
    "flipkart","paytm","freshworks","zoho","sprinklr",
    "cohere","anthropic","scale-ai","mistral","huggingface",
    "browserstack","hasura","supabase",
    "tiger-analytics","tredence","absolutdata","fractal",
    "grab","gojek","tokopedia","bukalapak","shopee",
    "byju","unacademy","vedantu","upgrad",
]

LEVER_SLUGS = [
    "slice","bharatpe","smallcase","setu","niyo","mswipe","happay",
    "databricks","canva","scaleai","carta","rippling","lattice",
    "mercury","deel","airtable","miro","loom","notion","linear",
    "clickup","asana","monday","pendo","mixpanel","heap","glean",
    "tredence","tiger-analytics","fractal","absolutdata",
    "zomato","dunzo","meesho","razorpay","cred",
    "kuda","mono","stitch",
    "cohere","anthropic","scale-ai","huggingface",
    "digitalocean","render","railway",
    "groww","zerodha","paytm","phonepe","freshworks",
    "sprinklr","zendesk","chargebee","postman",
    "browserstack","hasura","rudderstack",
]

ASHBY_SLUGS = [
    "razorpay","meesho","cred","zepto","swiggy","jupiter",
    "setu","cashfree","open","moneytap","yubi","credgenics",
    "browserstack","hasura","supabase","airbyte","fivetran",
    "prefect","lightdash","cube-dev",
    "cohere","mistral","together-ai","anyscale",
    "slope","ramp","brex","puzzle",
    "june","koala",
    "groww","zerodha","slice","niyo","bharatpe",
    "sprinklr","freshworks","chargebee",
]

WORKDAY_BOARDS = [
    ("Walmart",        "walmart",      "wd5",   "WalmartExternal"),
    ("Accenture",      "accenture",    "wd103", "AccentureCareers"),
    ("Barclays",       "barclays",     "wd3",   "External_Career_Site_Barclays"),
    ("Deloitte",       "deloitte",     "wd5",   "Deloitte_Careers"),
    ("EY",             "ey",           "wd5",   "EY_External"),
    ("PwC",            "pwc",          "wd3",   "Global_Campus_Experienced"),
    ("IBM",            "ibm",          "wd5",   "IBM"),
    ("Capgemini",      "capgemini",    "wd5",   "Capgemini"),
    ("Target",         "target",       "wd5",   "enterprise_careers"),
    ("Goldman Sachs",  "goldmansachs", "wd1",   "External_Career_Site"),
    ("Salesforce",     "salesforce",   "wd5",   "External_Career_Site"),
    ("Adobe",          "adobe",        "wd5",   "external_experienced"),
    ("PayPal",         "paypal",       "wd5",   "jobs"),
    ("Mastercard",     "mastercard",   "wd1",   "CorporateCareers"),
    ("Visa",           "visa",         "wd5",   "Visa"),
    ("Bosch",          "bosch",        "wd3",   "Bosch_Experienced_Positions"),
    ("SAP",            "sap",          "wd3",   "SAP"),
    ("Oracle",         "oracle",       "wd5",   "oracle-ext"),
    ("Micron",         "micron",       "wd5",   "External"),
    ("Honeywell",      "honeywell",    "wd5",   "Honeywell"),
    ("JPMorgan",       "jpmc",         "wd5",   "JPMCExternalSite"),
    ("Citi",           "citi",         "wd5",   "External"),
    ("HSBC",           "hsbc",         "wd3",   "external"),
    ("McKinsey",       "mckinsey",     "wd5",   "mck"),
    ("BCG",            "bcg",          "wd5",   "BCGCareers"),
    ("Bain",           "bain",         "wd5",   "BainCareers"),
]

SMARTRECRUITERS_SLUGS = [
    "Nykaa","PineLabs","PayUIndia","BharatPe","Ola","OlaMoney",
    "Policybazaar","IndiaMart","InfoEdge",
    "ThoughtWorks","EPAM","GlobalLogic","Sprinklr","Freshworks",
    "Zendesk","HCL-Technologies","Mphasis","LTIMindtree",
    "Booking.com","Delivery-Hero","OLX","Agoda","Klarna",
    "Randstad","ManpowerGroup","Gartner",
    "Revolut","Wise","Checkout.com","PayoneerInc",
    "WiproLimited","TCSiON","InfosysBPM",
]

# ─────────────────────────────────────────────────────────────────────────────
#  EMAIL / FILES
# ─────────────────────────────────────────────────────────────────────────────
EMAIL_SENDER       = os.environ.get("EMAIL_SENDER",    "")
EMAIL_PASSWORD     = os.environ.get("EMAIL_PASSWORD",  "")
EMAIL_RECIPIENT    = os.environ.get("EMAIL_RECIPIENT", "agastiabhishek@gmail.com")
ANTHROPIC_API_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")
SEEN_JOBS_FILE     = "seen_jobs.json"
EXCEL_FILE         = "jobs_tracker.xlsx"
RESUME_OUTPUT_DIR  = "resumes"

# ── Board category grouping for Excel demarcation ─────────────────────────────
BOARD_CATEGORY = {
    "Remotive":            "🌍 Remote Aggregator",
    "Jobicy":              "🌍 Remote Aggregator",
    "We Work Remotely":    "🌍 Remote Aggregator",
    "Remote OK":           "🌍 Remote Aggregator",
    "Working Nomads":      "🌍 Remote Aggregator",
    "Jobspresso":          "🌍 Remote Aggregator",
    "AngelList/Wellfound": "🌍 Remote Aggregator",
    "Surely":              "🌍 Remote Aggregator",
    "Greenhouse":          "🏢 ATS Board",
    "Lever":               "🏢 ATS Board",
    "Ashby":               "🏢 ATS Board",
    "Workday":             "🏢 ATS Board",
    "SmartRecruiters":     "🏢 ATS Board",
    "Instahyre":           "🇮🇳 India Portal",
    "Naukri":              "🇮🇳 India Portal",
    "iimjobs":             "🇮🇳 India Portal",
}

# Background tint pairs per category (even row, odd row) in Excel
CATEGORY_ROW_TINT = {
    "🌍 Remote Aggregator": ("E8F5E9", "D0EED4"),
    "🏢 ATS Board":         ("E3F2FD", "C5E3FA"),
    "🇮🇳 India Portal":    ("FFF8E1", "FDEFC4"),
}

# Divider bar colour per category in Excel
CATEGORY_DIVIDER_BG = {
    "🌍 Remote Aggregator": "2E7D32",
    "🏢 ATS Board":         "1565C0",
    "🇮🇳 India Portal":    "E65100",
}

# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def load_seen() -> set:
    if Path(SEEN_JOBS_FILE).exists():
        with open(SEEN_JOBS_FILE) as f:
            return set(json.load(f))
    return set()

def save_seen(seen: set) -> None:
    with open(SEEN_JOBS_FILE, "w") as f:
        json.dump(sorted(seen), f, indent=2)

def kw_match(text: str) -> bool:
    t = text.lower()
    return all(any(kw in t for kw in grp) for grp in KEYWORD_GROUPS)

def loc_match(text: str) -> bool:
    if not LOCATION_KEYWORDS:
        return True
    t = text.lower()
    return any(loc in t for loc in LOCATION_KEYWORDS)

def level_ok(text: str) -> bool:
    """Reject director/VP-level postings."""
    t = text.lower()
    if any(s in t for s in SENIOR_ONLY_KEYWORDS):
        return False
    return True

def html2txt(raw: str) -> str:
    return re.sub(r"<[^>]+>", " ", raw or "")

def job(uid, title, company, location, url, source, salary="", tags="", posted=""):
    return {
        "id":       uid,
        "title":    title,
        "company":  company,
        "location": location or "Remote / Not specified",
        "url":      url,
        "source":   source,
        "salary":   salary,
        "tags":     tags,
        "posted":   posted or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "scraped":  datetime.now(timezone.utc).strftime("%Y-%m-%d"),
    }

UA = "Mozilla/5.0 (compatible; JobBot/4.0)"

def _get(url, retries=3, **kw):
    for n in range(retries):
        try:
            r = requests.get(url, headers={"User-Agent": UA,
                                            "Accept": "application/json"},
                             timeout=20, **kw)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            print(f"      ⚠ GET[{n+1}] {url[:60]}: {e}")
            time.sleep(2 ** n)
    return None

def _get_rss(url, retries=3):
    for n in range(retries):
        try:
            r = requests.get(url, headers={"User-Agent": UA}, timeout=20)
            r.raise_for_status()
            return r.text
        except Exception as e:
            print(f"      ⚠ RSS[{n+1}] {url[:60]}: {e}")
            time.sleep(2 ** n)
    return None

def _post(url, retries=3, **kw):
    for n in range(retries):
        try:
            r = requests.post(url,
                              headers={"User-Agent": UA,
                                       "Content-Type": "application/json"},
                              timeout=20, **kw)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            print(f"      ⚠ POST[{n+1}] {url[:60]}: {e}")
            time.sleep(2 ** n)
    return None

def parse_rss_jobs(xml_text: str, source: str, prefix: str) -> list:
    """Generic RSS/Atom parser for job feeds."""
    out = []
    if not xml_text:
        return out
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return out
    ns = {"atom": "http://www.w3.org/2005/Atom"}
    items = root.findall(".//item") or root.findall(".//atom:entry", ns)
    for item in items:
        def t(tag): return (item.findtext(tag) or
                            item.findtext(f"atom:{tag}", namespaces=ns) or "")
        title    = t("title").strip()
        link     = t("link").strip() or t("id").strip()
        descr    = html2txt(t("description") or t("atom:summary") or "")
        pub      = t("pubDate") or t("published") or ""
        # category / tags
        cats = [c.text or "" for c in (item.findall("category") or [])]
        tags = ", ".join(c for c in cats if c)
        # location heuristic from title/description
        location = "Remote"
        for tok in ["India","Remote","Worldwide","Global","Anywhere",
                    "Bangalore","Mumbai","Delhi","Hyderabad","Pune"]:
            if tok.lower() in (title + descr).lower():
                location = tok; break
        uid = f"{prefix}_{re.sub(r'[^a-z0-9]','_', title.lower()[:30])}"
        if kw_match(f"{title} {descr}") and loc_match(f"{title} {location}") and level_ok(title):
            out.append(job(uid, title, source, location, link, source,
                           tags=tags, posted=pub[:10]))
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 1 — REMOTIVE  (all companies, free public API)
# ═══════════════════════════════════════════════════════════════════════════════
REMOTIVE_SEARCHES = [
    ("data engineer",      "software-dev"),
    ("data scientist",     "data"),
    ("analytics engineer", "data"),
    ("machine learning",   "data"),
    ("business analyst",   "data"),
    ("financial analyst",  "finance"),
    ("fraud analyst",      "finance"),
    ("product analyst",    "product"),
    ("mlops",              "devops-sysadmin"),
    ("llm engineer",       "software-dev"),
    ("data analyst",       "data"),
    ("bi analyst",         "data"),
    ("ai engineer",        "software-dev"),
]

def scrape_remotive() -> list:
    seen_ids, out = set(), []
    for search, cat in REMOTIVE_SEARCHES:
        data = _get("https://remotive.com/api/remote-jobs",
                    params={"search": search, "category": cat})
        if not data: continue
        for j in data.get("jobs", []):
            jid = f"rm_{j.get('id','')}"
            if jid in seen_ids: continue
            seen_ids.add(jid)
            title    = j.get("title", "")
            company  = j.get("company_name", "")
            location = j.get("candidate_required_location", "Worldwide")
            url      = j.get("url", "")
            descr    = html2txt(j.get("description", ""))
            salary   = j.get("salary", "")
            tags     = ", ".join(j.get("tags", []) or [])
            pub      = (j.get("publication_date") or "")[:10]
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url, "Remotive",
                               salary=str(salary), tags=tags, posted=pub))
        time.sleep(0.4)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    Remotive                  → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 2 — JOBICY  (all companies, free public API)
# ═══════════════════════════════════════════════════════════════════════════════
JOBICY_TAGS = [
    "data engineer","data scientist","analytics","machine learning",
    "business analyst","financial analyst","fraud detection",
    "mlops","databricks","python sql","data analyst","ai engineer",
]

def scrape_jobicy() -> list:
    seen_ids, out = set(), []
    for tag in JOBICY_TAGS:
        data = _get("https://jobicy.com/api/v2/remote-jobs",
                    params={"count": 50, "tag": tag})
        if not data: continue
        for j in data.get("jobs", []):
            jid = f"jc_{j.get('id','')}"
            if jid in seen_ids: continue
            seen_ids.add(jid)
            title    = j.get("jobTitle", "")
            company  = j.get("companyName", "")
            location = j.get("jobGeo", "Anywhere")
            url      = j.get("url", "")
            descr    = html2txt(j.get("jobDescription", ""))
            salary   = j.get("annualSalaryMin","")
            tags     = ", ".join(j.get("jobIndustry",[]) or [])
            pub      = (j.get("pubDate") or "")[:10]
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url, "Jobicy",
                               salary=str(salary), tags=tags, posted=pub))
        time.sleep(0.4)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    Jobicy                    → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 3 — WE WORK REMOTELY  (RSS feeds)
#  Categories relevant to Abhishek's profile
# ═══════════════════════════════════════════════════════════════════════════════
WWR_FEEDS = [
    ("https://weworkremotely.com/categories/remote-data-science-jobs.rss",   "WWR"),
    ("https://weworkremotely.com/categories/remote-programming-jobs.rss",    "WWR"),
    ("https://weworkremotely.com/categories/remote-finance-legal-jobs.rss",  "WWR"),
    ("https://weworkremotely.com/categories/remote-product-jobs.rss",        "WWR"),
    ("https://weworkremotely.com/remote-jobs.rss",                           "WWR"),
]

def scrape_weworkremotely() -> list:
    out = []
    for url, src in WWR_FEEDS:
        xml = _get_rss(url)
        out.extend(parse_rss_jobs(xml, "We Work Remotely", "wwr"))
        time.sleep(0.4)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    We Work Remotely (RSS)    → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 4 — REMOTE OK  (free public JSON API)
#  https://remoteok.com/api  — returns array, first element is metadata
# ═══════════════════════════════════════════════════════════════════════════════
REMOTEOK_TAGS = [
    "data-science","analytics","machine-learning","python","sql",
    "business-analyst","financial","fraud","mlops","ai","llm",
]

def scrape_remoteok() -> list:
    out = []
    seen_ids = set()
    for tag in REMOTEOK_TAGS:
        data = _get(f"https://remoteok.com/api?tag={tag}")
        if not data or not isinstance(data, list): continue
        for j in data[1:]:  # skip metadata entry
            jid = f"rok_{j.get('id','')}"
            if jid in seen_ids: continue
            seen_ids.add(jid)
            title    = j.get("position","")
            company  = j.get("company","")
            location = j.get("location","Worldwide")
            url      = j.get("url","") or f"https://remoteok.com/l/{j.get('slug','')}"
            descr    = html2txt(j.get("description",""))
            salary   = j.get("salary","")
            tags     = ", ".join(j.get("tags",[]) or [])
            pub      = (j.get("date",""))[:10]
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url, "Remote OK",
                               salary=str(salary), tags=tags, posted=pub))
        time.sleep(1)  # Remote OK rate-limits
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    Remote OK (JSON API)      → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 5 — WORKING NOMADS  (RSS)
# ═══════════════════════════════════════════════════════════════════════════════
WORKING_NOMADS_FEEDS = [
    "https://www.workingnomads.com/jobs?category=data-science&format=rss",
    "https://www.workingnomads.com/jobs?category=software-development&format=rss",
    "https://www.workingnomads.com/jobs?category=business&format=rss",
]

def scrape_workingnomads() -> list:
    out = []
    for url in WORKING_NOMADS_FEEDS:
        xml = _get_rss(url)
        out.extend(parse_rss_jobs(xml, "Working Nomads", "wn"))
        time.sleep(0.4)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    Working Nomads (RSS)      → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 6 — JOBSPRESSO  (RSS)
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_jobspresso() -> list:
    xml = _get_rss("https://jobspresso.co/feed/")
    out = parse_rss_jobs(xml, "Jobspresso", "jsp")
    print(f"    Jobspresso (RSS)          → {len(out)}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 7 — SURELY  (https://www.surely.co.in / surelyremote.com)
#  Surely is an aggregator that indexes 500+ job boards and company career pages.
#  It exposes a public JSON search API used by their website's job listing page.
#  We hit their search endpoint with multiple keyword queries, filtering for
#  remote / India roles matching Abhishek's profile.
# ═══════════════════════════════════════════════════════════════════════════════
SURELY_QUERIES = [
    "data scientist remote",
    "data analyst remote india",
    "business analyst remote india",
    "machine learning engineer remote",
    "data engineer remote",
    "analytics engineer remote",
    "ml engineer remote india",
    "bi analyst remote",
    "product analyst remote",
    "llm engineer remote",
    "ai engineer remote india",
]

def scrape_surely() -> list:
    out      = []
    seen_ids = set()

    # Surely exposes their job search via a public API used by their frontend.
    # Primary endpoint (v2 JSON):
    BASE_URL = "https://api.surely.co.in/api/v2/jobs/search"
    # Fallback endpoint if primary fails:
    FALLBACK  = "https://surelyremote.com/api/jobs"

    for q in SURELY_QUERIES:
        params = {
            "query":    q,
            "remote":   "true",
            "page":     1,
            "per_page": 30,
        }
        data = _get(BASE_URL, params=params)

        # Try alternate query key if first attempt fails
        if not data:
            data = _get(BASE_URL, params={"q": q, "remote": 1, "limit": 30})

        # Fallback to secondary endpoint
        if not data:
            data = _get(FALLBACK, params={"search": q, "remote": "true"})

        if not data:
            time.sleep(0.5)
            continue

        # Normalise: Surely may return a list or {"jobs": [...]} or {"results": [...]}
        if isinstance(data, list):
            items = data
        else:
            items = (data.get("jobs")
                     or data.get("results")
                     or data.get("data")
                     or [])

        for j in items:
            # Handle both flat dict and nested structures
            jid  = f"su_{j.get('id') or j.get('job_id') or j.get('slug','')}"
            if jid in seen_ids:
                continue
            seen_ids.add(jid)

            title    = (j.get("title")   or j.get("job_title")  or
                        j.get("position") or "")
            company  = (j.get("company") or j.get("company_name") or
                        (j.get("employer") or {}).get("name","") or "")
            location = (j.get("location") or j.get("job_location") or
                        j.get("remote_location") or "Remote")
            url      = (j.get("url")       or j.get("apply_url")   or
                        j.get("job_url")    or j.get("link")        or
                        f"https://surely.co.in/jobs/{j.get('id','')}")
            descr    = html2txt(j.get("description") or j.get("job_description") or "")
            salary   = str(j.get("salary") or j.get("salary_range") or "")
            tags     = ", ".join(j.get("skills") or j.get("tags") or [])
            pub      = (j.get("posted_at") or j.get("date_posted") or
                        j.get("created_at") or "")[:10]

            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url, "Surely",
                               salary=salary, tags=tags, posted=pub))
        time.sleep(0.5)

    final, ids = [], set()
    for j in out:
        if j["id"] not in ids:
            final.append(j); ids.add(j["id"])
    print(f"    Surely (aggregator)       → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 8 — WELLFOUND (AngelList)  — public JSON search
#  Uses the public search API (no auth required for basic listings)
# ═══════════════════════════════════════════════════════════════════════════════
WELLFOUND_ROLES = [
    "data scientist","data analyst","business analyst",
    "machine learning engineer","data engineer","ml engineer",
    "analytics engineer","product analyst","bi analyst",
]

def scrape_wellfound() -> list:
    out = []
    seen_ids = set()
    base = "https://wellfound.com/role/l"
    # Wellfound has public listing pages per role slug
    role_slugs = [
        "data-scientist","data-analyst","business-analyst",
        "machine-learning-engineer","data-engineer",
        "analytics-engineer","product-analyst",
    ]
    for slug in role_slugs:
        # Try their v2 talent API (public, no auth)
        data = _get(f"https://angel.co/api/v1/jobs?job_types=full-time&"
                    f"remote=true&roles[]={slug}&sort=recent")
        if not data or not isinstance(data.get("jobs"), list):
            time.sleep(0.4); continue
        for j in data["jobs"]:
            jid = f"wf_{j.get('id','')}"
            if jid in seen_ids: continue
            seen_ids.add(jid)
            title    = j.get("title","")
            company  = (j.get("startup") or {}).get("name","")
            location = j.get("remote_config","") or "Remote"
            url      = j.get("apply_url","") or f"https://wellfound.com/jobs/{j.get('id','')}"
            descr    = html2txt(j.get("description",""))
            salary   = j.get("compensation","")
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url,
                               "AngelList/Wellfound", salary=str(salary)))
        time.sleep(0.5)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    AngelList/Wellfound       → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 8 — GREENHOUSE  (per-board)
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_greenhouse_all() -> list:
    out = []
    for slug in GREENHOUSE_SLUGS:
        url  = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true"
        data = _get(url)
        if not data: time.sleep(0.3); continue
        for j in data.get("jobs", []):
            title    = j.get("title","")
            location = j.get("location",{}).get("name","")
            content  = html2txt(j.get("content",""))
            if (kw_match(f"{title} {content}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(
                    f"gh_{slug}_{j.get('id','')}",
                    title, slug.replace("-"," ").title(), location,
                    j.get("absolute_url",""), "Greenhouse",
                ))
        time.sleep(0.35)
    print(f"    Greenhouse ({len(GREENHOUSE_SLUGS)} boards)       → {len(out)}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 9 — LEVER  (per-board)
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_lever_all() -> list:
    out = []
    for slug in LEVER_SLUGS:
        data = _get(f"https://api.lever.co/v0/postings/{slug}?mode=json")
        if not data or not isinstance(data, list): time.sleep(0.3); continue
        for j in data:
            title    = j.get("text","")
            location = j.get("categories",{}).get("location","")
            descr    = html2txt(j.get("description","") + j.get("descriptionPlain",""))
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(
                    f"lv_{slug}_{j.get('id','')}",
                    title, slug.replace("-"," ").title(), location,
                    j.get("hostedUrl",""), "Lever",
                ))
        time.sleep(0.35)
    print(f"    Lever   ({len(LEVER_SLUGS)} boards)           → {len(out)}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 10 — ASHBY  (GraphQL per-board)
# ═══════════════════════════════════════════════════════════════════════════════
_ASHBY_URL = "https://jobs.ashbyhq.com/api/non-user-graphql?op=ApiJobBoardWithTeams"
_ASHBY_Q   = (
    "query ApiJobBoardWithTeams($organizationHostedJobsPageName:String!){"
    "jobBoard:jobBoardWithTeams(organizationHostedJobsPageName:$organizationHostedJobsPageName){"
    "jobPostings{id title locationName externalLink"
    " descriptionParts{descriptionHtml}}}}"
)

def scrape_ashby_all() -> list:
    out = []
    for slug in ASHBY_SLUGS:
        payload = {"operationName": "ApiJobBoardWithTeams",
                   "variables": {"organizationHostedJobsPageName": slug},
                   "query": _ASHBY_Q}
        data = _post(_ASHBY_URL, json=payload)
        if not data: time.sleep(0.3); continue
        postings = ((data.get("data") or {}).get("jobBoard") or {}).get("jobPostings",[]) or []
        for j in postings:
            title    = j.get("title","")
            location = j.get("locationName","")
            descr    = html2txt(" ".join(
                p.get("descriptionHtml","") for p in (j.get("descriptionParts") or [])
            ))
            url = j.get("externalLink") or f"https://jobs.ashbyhq.com/{slug}/{j.get('id','')}"
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(f"ab_{slug}_{j.get('id','')}", title,
                               slug.title(), location, url, "Ashby"))
        time.sleep(0.35)
    print(f"    Ashby   ({len(ASHBY_SLUGS)} boards)           → {len(out)}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 11 — WORKDAY  (per-tenant keyword POST)
# ═══════════════════════════════════════════════════════════════════════════════
_WD_SEARCH = (
    "data engineer data scientist analytics business analyst "
    "ml engineer financial analyst fraud analyst risk data analyst"
)

def scrape_workday_all() -> list:
    out = []
    for name, tenant, wd_ver, path in WORKDAY_BOARDS:
        base    = f"https://{tenant}.{wd_ver}.myworkdayjobs.com"
        api_url = f"{base}/wday/cxs/{tenant}/{path}/jobs"
        payload = {"appliedFacets": {}, "limit": 20, "offset": 0,
                   "searchText": _WD_SEARCH}
        data = _post(api_url, json=payload)
        if not data:
            for alt in ("wd5","wd3","wd1","wd103"):
                if alt == wd_ver: continue
                alt_url = f"https://{tenant}.{alt}.myworkdayjobs.com/wday/cxs/{tenant}/{path}/jobs"
                data = _post(alt_url, json=payload)
                if data: base = f"https://{tenant}.{alt}.myworkdayjobs.com"; break
        if not data: time.sleep(0.4); continue
        seen_t = set()
        for j in (data.get("jobPostings") or []):
            title    = j.get("title","")
            location = j.get("locationsText","")
            ext      = j.get("externalPath","")
            apply    = f"{base}/en-US/{path}{ext}" if ext else f"{base}/en-US/{path}"
            key      = title.lower().strip()
            if key in seen_t: continue
            seen_t.add(key)
            safe_id = re.sub(r"[^a-z0-9]","_", key)[:38]
            if (kw_match(title) and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(f"wd_{tenant}_{safe_id}", title, name,
                               location, apply, "Workday"))
        time.sleep(0.5)
    print(f"    Workday ({len(WORKDAY_BOARDS)} tenants)          → {len(out)}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 12 — SMARTRECRUITERS  (per-company, ?q= filter)
# ═══════════════════════════════════════════════════════════════════════════════
_SR_Q = (
    "data engineer OR data scientist OR analytics OR business analyst "
    "OR ml engineer OR fraud analyst OR financial analyst OR product analyst"
)

def scrape_smartrecruiters_all() -> list:
    out = []
    for slug in SMARTRECRUITERS_SLUGS:
        url    = f"https://api.smartrecruiters.com/v1/companies/{slug}/postings"
        params = {"q": _SR_Q, "limit": 100, "offset": 0}
        data   = _get(url, params=params)
        if not data: time.sleep(0.3); continue
        for j in (data.get("content") or []):
            title    = j.get("name","")
            lo       = j.get("location",{})
            parts    = [lo.get("city",""), lo.get("region",""),
                        lo.get("country",""), "Remote" if lo.get("remote") else ""]
            location = ", ".join(p for p in parts if p)
            pid      = j.get("id", j.get("uuid",""))
            apply    = f"https://jobs.smartrecruiters.com/{slug}/{pid}"
            if (kw_match(title) and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(f"sr_{slug}_{pid}", title,
                               slug.replace("-"," ").replace("."," ").title(),
                               location, apply, "SmartRecruiters"))
        time.sleep(0.35)
    print(f"    SmartRecruiters ({len(SMARTRECRUITERS_SLUGS)} cos)    → {len(out)}")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 13 — INSTAHYRE  (India — public search API)
# ═══════════════════════════════════════════════════════════════════════════════
INSTAHYRE_QUERIES = [
    "data scientist","data analyst","business analyst",
    "machine learning","data engineer","analytics",
]

def scrape_instahyre() -> list:
    out = []
    seen_ids = set()
    for q in INSTAHYRE_QUERIES:
        try:
            r = requests.get(
                "https://www.instahyre.com/api/v1/opportunity/",
                params={"search": q, "remote": "true",
                        "experience_min": 1, "experience_max": 4},
                headers={"User-Agent": UA, "Accept": "application/json"},
                timeout=20
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"      ⚠ Instahyre [{q}]: {e}"); time.sleep(0.5); continue
        for j in (data if isinstance(data, list) else data.get("results", [])):
            jid = f"ih_{j.get('id','')}"
            if jid in seen_ids: continue
            seen_ids.add(jid)
            title   = j.get("role","") or j.get("title","")
            company = (j.get("company") or {}).get("name","") or j.get("company_name","")
            location = j.get("location","India") or "India (Remote)"
            url      = j.get("job_url","") or f"https://www.instahyre.com/candidate/opportunities/{j.get('id','')}"
            descr    = j.get("description","")
            salary   = j.get("salary","") or ""
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url,
                               "Instahyre", salary=str(salary)))
        time.sleep(0.6)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    Instahyre (India)         → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 14 — NAUKRI  (India — public job search via their search API)
# ═══════════════════════════════════════════════════════════════════════════════
NAUKRI_SEARCHES = [
    "data scientist remote","data analyst remote india",
    "business analyst remote india","machine learning engineer remote",
    "data engineer remote india","analytics remote india",
]

def scrape_naukri() -> list:
    out = []
    seen_ids = set()
    headers = {
        "User-Agent":   UA,
        "appid":        "109",
        "systemid":     "109",
        "Accept":       "application/json",
    }
    for q in NAUKRI_SEARCHES:
        try:
            r = requests.get(
                "https://www.naukri.com/jobapi/v3/search",
                params={"noOfResults": 20, "urlType": "search_by_keyword",
                        "searchType": "adv", "keyword": q,
                        "pageNo": 1, "seoKey": q.replace(" ","-"),
                        "src": "jobsearchDesk", "latLong": ""},
                headers=headers, timeout=25
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"      ⚠ Naukri [{q}]: {e}"); time.sleep(0.5); continue
        for j in (data.get("jobDetails") or []):
            jid = f"nk_{j.get('jobId','')}"
            if jid in seen_ids: continue
            seen_ids.add(jid)
            title    = j.get("title","")
            company  = j.get("companyName","")
            location = ", ".join(j.get("placeholders",[{}])[0:1] and
                                 [p.get("label","") for p in j.get("placeholders",[])
                                  if p.get("type") == "location"]) or "India"
            salary   = ", ".join([p.get("label","") for p in j.get("placeholders",[])
                                  if p.get("type") == "salary"])
            url      = j.get("jdURL","") or f"https://www.naukri.com{j.get('jobId','')}"
            descr    = j.get("jobDescription","")
            if (kw_match(f"{title} {descr}") and
                    loc_match(f"{title} {location}") and level_ok(title)):
                out.append(job(jid, title, company, location, url,
                               "Naukri", salary=salary))
        time.sleep(0.7)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    Naukri (India)            → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  SOURCE 15 — iimjobs  (Premium India roles — RSS)
# ═══════════════════════════════════════════════════════════════════════════════
def scrape_iimjobs() -> list:
    feeds = [
        "https://www.iimjobs.com/j/data-science-analytics,236/data-analyst,7.html?rss=1",
        "https://www.iimjobs.com/j/business-analyst,16.html?rss=1",
    ]
    out = []
    for url in feeds:
        xml = _get_rss(url)
        out.extend(parse_rss_jobs(xml, "iimjobs", "iim"))
        time.sleep(0.4)
    final, ids = [], set()
    for j in out:
        if j["id"] not in ids: final.append(j); ids.add(j["id"])
    print(f"    iimjobs (India premium)   → {len(final)}")
    return final

# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

# Color palette
COL_HEADER_BG  = "1A2F5A"   # dark navy
COL_HEADER_FG  = "FFFFFF"
COL_ALT_ROW    = "EEF3FA"   # light blue
COL_STRIPE     = "FFFFFF"

# Source colour mapping for pills
SOURCE_COLORS = {
    "Remotive":            "E84040",
    "Jobicy":              "F07020",
    "We Work Remotely":    "2563EB",
    "Remote OK":           "059669",
    "Working Nomads":      "7C3AED",
    "Jobspresso":          "DB2777",
    "Surely":              "0891B2",
    "AngelList/Wellfound": "EA580C",
    "Greenhouse":          "1F6FBF",
    "Lever":               "1A6B3C",
    "Ashby":               "5B3090",
    "Workday":             "B84A0D",
    "SmartRecruiters":     "1A3566",
    "Instahyre":           "0F766E",
    "Naukri":              "B45309",
    "iimjobs":             "9D174D",
}

EXCEL_COLUMNS = [
    ("Board Category",   22),
    ("Job Title",        35),
    ("Company",          25),
    ("Location",         22),
    ("Source",           18),
    ("Salary / Range",   20),
    ("Tags / Skills",    35),
    ("Date Posted",      14),
    ("Date Scraped",     14),
    ("Apply Link",       20),
    ("Status",           16),
    ("Notes",            30),
    ("Match Score",      14),
    ("Tailored Resume",  22),   # Col N — ATS-optimised PDF path / hyperlink
]

def _match_score(j: dict) -> str:
    """Simple heuristic match score based on profile alignment."""
    title = j["title"].lower()
    tags  = j["tags"].lower()
    score = 0
    # Core titles
    for kw in ["data scientist","data analyst","analytics","ml engineer",
               "machine learning","business analyst"]:
        if kw in title: score += 30; break
    # Premium tech stack match
    for kw in ["python","sql","pyspark","databricks","azure","power bi",
               "xgboost","mlflow","langchain","rag","llm"]:
        if kw in tags or kw in title: score += 5
    score = min(score, 100)
    if score >= 80:   return "⭐⭐⭐ Excellent"
    elif score >= 55: return "⭐⭐ Good"
    elif score >= 30: return "⭐ Fair"
    else:             return "Review"

# ═══════════════════════════════════════════════════════════════════════════════
#  RESUME GENERATION  —  Claude API + ReportLab
#  For every new job, call Claude to rewrite the base resume to match the JD,
#  then render it as a clean 1-page PDF using ReportLab.
# ═══════════════════════════════════════════════════════════════════════════════

# ── Base resume text (structured for easy rewriting) ──────────────────────────
BASE_RESUME = """
NAME: Abhishek Agasti
TITLE: Analyst — Data Science | Business Analyst | Cloud Data Platforms
CONTACT: agastiabhishek@gmail.com | 8473093618 | India | linkedin.com/in/abhishek-agasti/

SUMMARY:
Data Science professional with hands-on experience in machine learning, deep learning,
NLP and Generative AI (LLMs). Skilled in Python, SQL, PySpark and Azure/Databricks to
build, validate and deploy predictive models and RAG-based solutions. Delivered 16+
industry-grade projects (forecasting, pricing optimisation, fraud detection, RAG BI)
with measurable business impact including 62% improvement in forecasting accuracy.
Experienced in model validation, explainability (SHAP) and MLOps practices to support
compliant, production-ready deployments and stakeholder communication.

EXPERIENCE:

Data Science & AI Trainee | Intellipaat — IIT Roorkee | Feb 2025 – Feb 2026
- Delivered 25+ AI/ML projects in forecasting, NLP, and cloud engineering leveraging Python, Azure, and SQL.
- Addressed complex business problems across 60+ datasets using advanced statistical modelling, achieving 20% improvement in predictive accuracy within 3 months.
- Engineered 100,000+ forecast rows for air quality prediction using PySpark and ARIMA models, improving city-level planning.
- Developed 6 pricing scenarios with revenue margin analysis using Python (Pandas, NumPy, Matplotlib, Seaborn), saving ~8 hrs/week.
- Analysed 10+ users and 10+ events with advanced SQL, identifying key drivers for 15% MRR growth.
- Applied MLflow experiment tracking and SHAP explainability to support deployment decisions and communicate risk/performance to stakeholders.

Business Development & BI Executive II | Asian Paints | May 2024 – Jun 2025
- Analysed regional sales data to identify growth opportunities, driving a 20% revenue increase in key segments within 4 months.
- Designed Excel-based analytics dashboards tracking 15 KPIs across product lines, enabling the regional sales manager to optimise lead conversion pipeline.
- Advised 50+ construction companies by translating complex data into actionable recommendations, driving GTM strategy.
- Developed Tableau dashboards for 500K+ records, enhancing business intelligence across 7 departments.
- Designed BI dashboards to track 15 KPIs, enabling product managers to optimise lead conversion by 15%.
- Engineered an A/B test framework to determine price-effect size across 10+ configurations.

GTM Strategy & Revenue Growth Intern | Asian Paints | May 2023 – Oct 2023
- Standardised 20+ business KPIs across the organisation by performing end-to-end EDA.
- Analysed 500K records using SQL, generating insights that improved business intelligence across 4 departments.
- Engineered 10 pricing scenarios using Python (Pandas, NumPy), saving ~15 hrs/week.
- Developed 15 AI/ML projects using Python and Azure, enhancing forecasting accuracy by 10%.

Data Analyst | Webx Technologies Pvt Ltd | Jul 2021 – Jun 2022
- Engineered and maintained BI solutions using Python, SQL, Power BI, Pandas, and PostgreSQL, improving data processing efficiency by 10%.
- Analysed 500K records using SQL, improving business intelligence across multiple departments.
- Applied model risk management principles and project management skills to ensure accuracy of data science models.

PROJECTS:

Go-to-Market (GTM) Performance Dashboard
- Engineered a GTM analytics framework using Python & SQL for 9,994 transactions across 10 industries, yielding $2.3M pipeline and identifying $322K revenue risk.
- Automated weekly SQL report refreshes saving 260+ hrs/year and flagging stalled deals >30 days.

NIFTY 50 Market Forecasting & Financial Intelligence
- Architected a multi-model equity framework (ARIMA, SARIMA, Prophet), reducing directional error by 30%.
- Developed 15+ technical indicators achieving 60%+ directional accuracy for RSI signals over 2 years.

Real-Time Fraud Detection System
- Architected an ensemble fraud detection pipeline (Isolation Forest + XGBoost + SMOTE) achieving strong AUC-ROC and enabling 50ms real-time authorisation.
- Deployed using Docker and Redis with Prometheus/Grafana for real-time monitoring of latency and data drift.
- Leveraged SHAP explainability for auditable, compliance-ready decisions.

Dynamic Pricing Optimisation & Demand Forecasting
- Engineered ensemble forecasting pipeline (Prophet + XGBoost), cutting RMSE by 62% for 4-week inventory planning.
- Engineered A/B test framework (power=0.8, α=0.05) with MLflow tracking across 20+ configurations.

SaaS Product Analytics Dashboard
- Architected a full-stack SaaS analytics pipeline, driving MRR growth from $47K to $196K (318%).
- Led cohort analysis and A/B testing achieving 6.8pp conversion lift (32% → 38.8%).

RAG Business Intelligence Pipeline (BizIntel)
- Deployed a scalable RAG application using Python, LangChain, FAISS, and FastAPI processing 15+ user queries daily.
- Automated data retrieval reducing report generation time by 12% and saving analysts 4 hours/week.

AQI Forecasting — Azure Data Pipeline
- Engineered an Azure data pipeline using PySpark and Databricks processing 500 GB daily.
- Built end-to-end Azure pipeline using Medallion Architecture (Databricks/PySpark, Delta Lake, Synapse Analytics, Data Factory).

EDUCATION:
Post Graduate Diploma in Data Engineering — IIT Jodhpur (Jan 2025 – Dec 2025) | CGPA: 7.8
MBA Advanced Project Management — NICMAR Pune (Aug 2022 – Apr 2024) | CGPA: 8.0
B.E. Civil Engineering — Assam Science and Technology University (Jun 2017 – May 2021) | 75.39%

CERTIFICATIONS:
- Databricks Certified Associate Developer — Apache Spark 3.0
- Executive Post Graduate Certification in Data Science & AI — Feb 2025 – Feb 2026

SKILLS:
Programming: Python | SQL | PySpark
ML & AI: Scikit-learn | XGBoost | TensorFlow | Transformers | LLMs | LangChain | FAISS | Generative AI | Prompt Engineering
Modelling: Model validation | Explainable AI (SHAP) | Predictive modelling | Time series (ARIMA, SARIMA, Prophet)
MLOps: MLflow | Docker | Flask | GitHub Actions CI/CD | Prometheus/Grafana
Cloud & Big Data: Azure | Databricks | Delta Lake | Synapse Analytics | Apache Spark
BI & Visualisation: Power BI | PostgreSQL | Tableau
Dev Tools: Microsoft SQL Server | Git | GitHub | Jupyter Notebook | VS Code
"""

# ── ATS keyword extraction & scoring helpers ──────────────────────────────────

def _extract_jd_keywords(job_title: str, job_tags: str, jd_text: str) -> list:
    """
    Pull every meaningful keyword/phrase from JD title + tags + description.
    Returns a deduplicated lowercase list ordered by likely ATS weight.
    """
    raw = f"{job_title} {job_tags} {jd_text}".lower()

    # Remove HTML tags if any leaked through
    raw = re.sub(r"<[^>]+>", " ", raw)

    # Known high-value tech & role tokens (extend as needed)
    priority_tokens = [
        # Role titles
        "data scientist","data science","data analyst","data analytics",
        "business analyst","analytics engineer","ml engineer","ai engineer",
        "machine learning engineer","applied scientist","decision scientist",
        "product analyst","bi analyst","bi developer","business intelligence",
        "llm engineer","nlp engineer","mlops engineer","data engineer",
        "financial analyst","fraud analyst","risk analyst","pricing analyst",
        # Core tech
        "python","sql","pyspark","spark","databricks","azure","aws","gcp",
        "power bi","tableau","looker","dbt","snowflake","bigquery","redshift",
        "airflow","kafka","delta lake","synapse","data factory",
        # ML / AI
        "machine learning","deep learning","nlp","natural language processing",
        "generative ai","large language models","llm","llms","transformers",
        "rag","retrieval augmented generation","langchain","faiss","openai",
        "xgboost","scikit-learn","tensorflow","pytorch","keras",
        "mlflow","mlops","model deployment","model validation","shap",
        "a/b testing","experiment tracking","feature engineering",
        # Statistics / modelling
        "arima","sarima","prophet","time series","forecasting",
        "regression","classification","clustering","anomaly detection",
        "statistical modelling","predictive modelling","causal inference",
        # Data engineering
        "etl","elt","data pipeline","data warehouse","data lake","medallion",
        "delta","parquet","api","rest api","fastapi","flask","docker",
        "kubernetes","ci/cd","github actions","git",
        # Business / soft
        "stakeholder","cross-functional","communication","insights",
        "dashboard","kpi","metrics","revenue","growth","optimisation",
        "a/b","cohort analysis","funnel","retention","conversion",
        "fraud detection","risk management","compliance","governance",
    ]

    found = []
    for tok in priority_tokens:
        if tok in raw:
            found.append(tok)

    # Also extract 1-2 word noun phrases from the raw text that look technical
    words = re.findall(r"\b[a-z][a-z0-9+#.-]{2,}\b", raw)
    for w in words:
        if (w not in found and len(w) >= 4 and
                w not in {"with","that","this","have","from","into","will",
                          "your","their","been","also","more","over","under",
                          "each","some","able","than","then","them","they",
                          "when","where","what","which","while","after","about",
                          "using","based","across","within","through","between"}):
            found.append(w)

    # Deduplicate preserving order
    seen_kw, unique = set(), []
    for k in found:
        if k not in seen_kw:
            seen_kw.add(k); unique.append(k)

    return unique[:120]   # cap at 120 distinct keywords


def _ats_score(resume_text: str, keywords: list) -> tuple:
    """
    Measure what % of JD keywords appear in the resume text.
    Returns (score_float 0-1, missing_keywords list).
    """
    rt_lower = resume_text.lower()
    matched  = [k for k in keywords if k in rt_lower]
    missing  = [k for k in keywords if k not in rt_lower]
    score    = len(matched) / len(keywords) if keywords else 1.0
    return score, missing


# ── Claude API call — 95%+ ATS keyword injection ──────────────────────────────
_SYSTEM_ATS = """You are a senior ATS (Applicant Tracking System) resume optimisation specialist.
Your ONLY task is to rewrite a candidate's resume to achieve a MINIMUM 95% keyword match
against a specific job description, while maintaining 100% factual accuracy.

STRICT RULES:
1. STRUCTURE IS FIXED — same section order, same jobs, same projects, same dates.
   Never add new employers, projects, certifications, or qualifications.
2. KEYWORD INJECTION — every high-priority JD keyword must appear naturally in the
   summary, bullets, or skills section. Do not keyword-stuff; weave terms into
   achievement-focused sentences.
3. VERB ALIGNMENT — use the exact verb tense and phrasing the JD uses
   (e.g. if JD says "build and deploy", use "built and deployed").
4. SUMMARY — 4-5 sentences. Must mention the job title, top 5 hard skills from the JD,
   and two quantified achievements.
5. BULLETS — each bullet must be STAR-format: action verb + tool/technology + outcome.
   Lead with JD keywords. Include numbers where possible.
6. SKILLS SECTION — list skills grouped by category. Include ALL technical skills
   mentioned in the JD that the candidate possesses. Do NOT invent skills.
7. EDUCATION & CERTIFICATIONS — kept verbatim; do not reword.
8. OUTPUT FORMAT — return ONLY a valid JSON object. No markdown fences, no prose.

JSON SCHEMA (follow exactly):
{
  "summary": "string — 4-5 sentences, JD-keyword-rich",
  "education": [
    {"degree": "string", "institution": "string", "period": "string", "grade": "string"}
  ],
  "certifications": ["string", ...],
  "experience": [
    {
      "role": "string",
      "company": "string",
      "period": "string",
      "bullets": ["string", ...]
    }
  ],
  "projects": [
    {
      "name": "string",
      "bullets": ["string", ...]
    }
  ],
  "skills": [
    {"category": "string", "items": "string"}
  ]
}"""


def _call_claude_resume(job_title: str, company: str, jd_text: str,
                        job_tags: str, keywords: list,
                        missing_from_prev: list = None) -> dict:
    """
    Call Claude API to rewrite the base resume for a specific JD.
    keywords        — extracted JD keyword list for injection guidance
    missing_from_prev — keywords still missing after a previous attempt (retry)
    Returns parsed dict or {} on failure.
    """
    if not ANTHROPIC_API_KEY:
        return {}

    kw_str     = ", ".join(keywords[:80])
    missing_str = ""
    if missing_from_prev:
        missing_str = (
            f"\n\nCRITICAL — STILL MISSING FROM PREVIOUS ATTEMPT "
            f"(you MUST include ALL of these):\n"
            + ", ".join(missing_from_prev[:40])
        )

    user_prompt = (
        f"TARGET ROLE: {job_title} at {company}\n"
        f"JD FULL TEXT: {jd_text[:2000]}\n"
        f"ALL JD KEYWORDS TO EMBED (priority order): {kw_str}"
        f"{missing_str}\n\n"
        f"BASE RESUME TO REWRITE:\n{BASE_RESUME}\n\n"
        "Rewrite the resume now. Every keyword in the list must appear at least once "
        "in summary, bullets, or skills. Use exact JD terminology. "
        "Return ONLY the JSON object."
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key":         ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type":      "application/json",
            },
            json={
                "model":      "claude-sonnet-4-20250514",
                "max_tokens": 4000,
                "system":     _SYSTEM_ATS,
                "messages":   [{"role": "user", "content": user_prompt}],
            },
            timeout=90,
        )
        resp.raise_for_status()
        raw = resp.json()["content"][0]["text"].strip()
        raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        print(f"      ⚠ Claude resume API error: {e}")
        return {}


def _rewritten_to_text(d: dict) -> str:
    """Flatten all rewritten content into a single string for ATS scoring."""
    parts = [d.get("summary","")]
    for exp in (d.get("experience") or []):
        parts.append(exp.get("role",""))
        parts += (exp.get("bullets") or [])
    for proj in (d.get("projects") or []):
        parts.append(proj.get("name",""))
        parts += (proj.get("bullets") or [])
    for sk in (d.get("skills") or []):
        parts.append(sk.get("items",""))
    return " ".join(parts)


# ── ReportLab PDF builder ──────────────────────────────────────────────────────
_NAVY    = colors.HexColor("#1A2F5A")
_BLUE    = colors.HexColor("#1565C0")
_GRAY    = colors.HexColor("#555555")
_LGRAY   = colors.HexColor("#E8EEF5")
_BLACK   = colors.black
_WHITE   = colors.white

def _build_resume_pdf(job: dict, rewritten: dict, pdf_path: str,
                      ats_score: float = 0.0) -> None:
    """
    Render an ATS-friendly PDF resume.
    Section order (matching request):
      1. Header
      2. Professional Summary
      3. Education & Certifications   ← moved up
      4. Technical Skills
      5. Professional Experience
      6. Key Projects
    """
    if not REPORTLAB_AVAILABLE:
        return

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=letter,
        leftMargin=0.55*inch, rightMargin=0.55*inch,
        topMargin=0.45*inch,  bottomMargin=0.45*inch,
    )

    normal_font = "Helvetica"
    bold_font   = "Helvetica-Bold"
    PAGE_W      = letter[0] - 1.1*inch

    # ── Style definitions ──────────────────────────────────────────────────────
    s_name    = ParagraphStyle("rname",    fontName=bold_font,   fontSize=16,
                                textColor=_NAVY, spaceAfter=1, leading=19)
    s_title   = ParagraphStyle("rtitle",   fontName=normal_font, fontSize=9.5,
                                textColor=_BLUE, spaceAfter=2, leading=12)
    s_contact = ParagraphStyle("rcontact", fontName=normal_font, fontSize=8.2,
                                textColor=_GRAY, spaceAfter=5, leading=11)
    s_section = ParagraphStyle("rsect",    fontName=bold_font,   fontSize=10,
                                textColor=_NAVY, spaceBefore=6, spaceAfter=1,
                                leading=13)
    s_role    = ParagraphStyle("rrole",    fontName=bold_font,   fontSize=9.2,
                                textColor=_BLACK, spaceBefore=4, spaceAfter=0,
                                leading=12)
    s_period  = ParagraphStyle("rperiod",  fontName=normal_font, fontSize=8.2,
                                textColor=_GRAY, spaceBefore=0, spaceAfter=1,
                                leading=11)
    s_bullet  = ParagraphStyle("rbullet",  fontName=normal_font, fontSize=8.6,
                                textColor=_BLACK, spaceBefore=0, spaceAfter=1,
                                leading=11.2, leftIndent=12)
    s_summary = ParagraphStyle("rsum",     fontName=normal_font, fontSize=8.6,
                                textColor=_BLACK, spaceAfter=3, leading=12)
    s_edu     = ParagraphStyle("redu",     fontName=normal_font, fontSize=8.6,
                                textColor=_BLACK, spaceAfter=1, leading=11)
    s_cert    = ParagraphStyle("rcert",    fontName=normal_font, fontSize=8.6,
                                textColor=_BLACK, spaceAfter=1, leading=11)
    s_footer  = ParagraphStyle("rfoot",    fontName="Helvetica-Oblique",
                                fontSize=7, textColor=_GRAY,
                                alignment=TA_CENTER, spaceAfter=0)

    def hr(thick=0.8):
        return HRFlowable(width="100%", thickness=thick,
                          color=_NAVY, spaceAfter=2, spaceBefore=1)
    def sec(title):
        return [Paragraph(title.upper(), s_section), hr()]
    def bullet_p(text):
        return Paragraph(f"\u2022  {text}", s_bullet)

    story = []

    # ── 1. Header ──────────────────────────────────────────────────────────────
    story.append(Paragraph("ABHISHEK AGASTI", s_name))
    story.append(Paragraph(
        f"Data Science Analyst  \u00B7  Business Analyst  \u00B7  "
        f"Cloud &amp; ML Platforms", s_title))
    story.append(Paragraph(
        f"<b>Application for: {job['title']} @ {job['company']}</b>",
        ParagraphStyle("rappfor", fontName="Helvetica-Bold", fontSize=9.5,
                       textColor=_BLUE, spaceAfter=2, leading=12)))
    story.append(Paragraph(
        "agastiabhishek@gmail.com  \u00B7  +91-8473093618  \u00B7  India  \u00B7  "
        "<a href=\"https://www.linkedin.com/in/abhishek-agasti/\"><u>LinkedIn</u></a>  \u00B7  "
        "<a href=\"https://github.com/abhishekagasti23\"><u>GitHub</u></a>",
        s_contact))
    story.append(HRFlowable(width="100%", thickness=1.5, color=_NAVY,
                             spaceAfter=3, spaceBefore=0))

    # ── 2. Professional Summary ────────────────────────────────────────────────
    story.extend(sec("Professional Summary"))
    summary_text = (
        rewritten.get("summary") or
        "Data Science professional with 3+ years of experience in machine learning, "
        "NLP, and Generative AI. Skilled in Python, SQL, PySpark, and "
        "Azure/Databricks to build, validate, and deploy predictive models and "
        "RAG-based solutions. Delivered 16+ industry-grade projects across "
        "forecasting, fraud detection, and pricing optimisation — achieving 62% "
        "improvement in RMSE and 20% improvement in predictive accuracy."
    )
    story.append(Paragraph(summary_text, s_summary))

    # ── 3. Education & Certifications (SECTION 2 as requested) ─────────────────
    story.extend(sec("Education & Certifications"))

    edu_data = rewritten.get("education") or [
        {"degree": "Post Graduate Diploma in Data Engineering",
         "institution": "IIT Jodhpur", "period": "Jan 2025 – Dec 2025", "grade": "CGPA: 7.8"},
        {"degree": "MBA — Advanced Project Management",
         "institution": "NICMAR, Pune", "period": "Aug 2022 – Apr 2024", "grade": "CGPA: 8.0"},
        {"degree": "B.E. Civil Engineering",
         "institution": "Assam Science and Technology University",
         "period": "Jun 2017 – May 2021", "grade": "75.39%"},
    ]
    for edu in edu_data:
        story.append(Paragraph(
            f"<b>{edu.get('degree','')}</b>  —  {edu.get('institution','')}  "
            f"|  {edu.get('period','')}  |  {edu.get('grade','')}",
            s_edu))

    certs = rewritten.get("certifications") or [
        "Databricks Certified Associate Developer — Apache Spark 3.0",
        "Executive Post Graduate Certification in Data Science & AI  (Feb 2025 – Feb 2026)",
    ]
    for cert in certs:
        story.append(Paragraph(f"\u2022  {cert}", s_cert))

    # ── 4. Technical Skills ────────────────────────────────────────────────────
    story.extend(sec("Technical Skills"))

    skills_data = rewritten.get("skills") or [
        {"category": "Programming",   "items": "Python  |  SQL  |  PySpark"},
        {"category": "ML & AI",       "items": "Scikit-learn  |  XGBoost  |  TensorFlow  |  Transformers  |  LLMs  |  LangChain  |  FAISS  |  Generative AI  |  Prompt Engineering"},
        {"category": "Modelling",     "items": "SHAP  |  Predictive Modelling  |  Time Series (ARIMA, SARIMA, Prophet)  |  Model Validation  |  Explainable AI"},
        {"category": "MLOps",         "items": "MLflow  |  Docker  |  Flask  |  GitHub Actions CI/CD  |  Prometheus  |  Grafana"},
        {"category": "Cloud & Data",  "items": "Azure  |  Databricks  |  Delta Lake  |  Synapse Analytics  |  Data Factory  |  Apache Spark"},
        {"category": "BI & Viz",      "items": "Power BI  |  Tableau  |  PostgreSQL  |  BigQuery"},
        {"category": "Dev Tools",     "items": "Git  |  GitHub  |  Jupyter Notebook  |  VS Code  |  Microsoft SQL Server"},
    ]
    for sk in skills_data:
        cat  = sk.get("category") or sk.get("cat","")
        vals = sk.get("items") or sk.get("values","")
        tbl  = Table(
            [[Paragraph(f"<b>{cat}</b>", ParagraphStyle(
                  "skl", fontName=bold_font, fontSize=8.5,
                  textColor=_NAVY, leading=11)),
              Paragraph(vals, ParagraphStyle(
                  "skv", fontName=normal_font, fontSize=8.5,
                  textColor=_BLACK, leading=11))]],
            colWidths=[1.15*inch, PAGE_W - 1.15*inch],
        )
        tbl.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ("TOPPADDING",    (0,0), (-1,-1), 1),
            ("BOTTOMPADDING", (0,0), (-1,-1), 1),
        ]))
        story.append(tbl)

    # ── 5. Professional Experience ─────────────────────────────────────────────
    story.extend(sec("Professional Experience"))

    experience_data = rewritten.get("experience") or [
        {
            "role": "Data Science & AI Trainee",
            "company": "Intellipaat — IIT Roorkee",
            "period": "Feb 2025 – Feb 2026",
            "bullets": [
                "Delivered 25+ AI/ML projects in forecasting, NLP, and cloud data engineering leveraging Python, Azure, and SQL.",
                "Addressed business problems across 60+ datasets via advanced statistical modelling — achieving 20% improvement in predictive accuracy within 3 months.",
                "Engineered 100,000+ forecast rows for air quality prediction using PySpark and ARIMA, improving city-level planning.",
                "Applied MLflow experiment tracking and SHAP explainability to support model deployment decisions and stakeholder risk communication.",
                "Developed pricing scenarios with revenue margin analysis using Pandas/NumPy/Seaborn, saving ~8 hrs/week.",
            ]
        },
        {
            "role": "Business Development & BI Executive II",
            "company": "Asian Paints",
            "period": "May 2024 – Jun 2025",
            "bullets": [
                "Analysed regional sales data to identify growth opportunities — drove 20% revenue increase in key segments within 4 months.",
                "Designed analytics dashboards tracking 15 KPIs across product lines, enabling the sales manager to optimise lead conversion pipeline.",
                "Developed Tableau dashboards for 500K+ records, enhancing business intelligence across 7 departments.",
                "Engineered A/B test framework to determine price-effect size across 10+ configurations; reported findings to regional leadership.",
            ]
        },
        {
            "role": "GTM Strategy & Revenue Growth Intern",
            "company": "Asian Paints",
            "period": "May 2023 – Oct 2023",
            "bullets": [
                "Standardised 20+ business KPIs via end-to-end EDA and cross-functional collaboration.",
                "Analysed 500K records using SQL, generating insights that improved BI across 4 departments.",
                "Engineered 10 pricing scenarios using Python (Pandas, NumPy), saving ~15 hrs/week on revenue margin analysis.",
            ]
        },
        {
            "role": "Data Analyst",
            "company": "Webx Technologies Pvt Ltd",
            "period": "Jul 2021 – Jun 2022",
            "bullets": [
                "Engineered BI solutions using Python, SQL, Power BI, Pandas, and PostgreSQL — improving data processing efficiency by 10%.",
                "Applied model risk management principles to ensure accuracy and reliability of analytics pipelines.",
            ]
        },
    ]

    for exp in experience_data:
        story.append(Paragraph(
            f"<b>{exp.get('role','')}</b>  \u2022  {exp.get('company','')}",
            s_role))
        story.append(Paragraph(exp.get("period",""), s_period))
        for b in (exp.get("bullets") or []):
            story.append(bullet_p(b))

    # ── 6. Key Projects ────────────────────────────────────────────────────────
    story.extend(sec("Key Projects"))

    projects_data = rewritten.get("projects") or [
        {"name": "Real-Time Fraud Detection  (Financial ML & Risk Analytics)",
         "bullets": [
             "Architected ensemble pipeline (Isolation Forest + XGBoost + SMOTE); strong AUC-ROC, 50ms real-time authorisation.",
             "SHAP explainability for compliance-ready auditable decisions; Docker + Redis deployment with Prometheus/Grafana monitoring.",
         ]},
        {"name": "Dynamic Pricing Optimisation & Demand Forecasting",
         "bullets": [
             "Prophet + XGBoost ensemble cutting RMSE by 62% (92K→35K) for 4-week inventory planning.",
             "A/B test framework (power=0.8, α=0.05) with MLflow across 20+ configurations; modular real-time pricing API.",
         ]},
        {"name": "RAG Business Intelligence Pipeline (BizIntel)",
         "bullets": [
             "LangChain + FAISS + FastAPI RAG app processing 15+ user queries daily; reduced report generation time by 12%.",
         ]},
        {"name": "AQI Forecasting — Azure Data Pipeline",
         "bullets": [
             "500 GB/day PySpark/Databricks pipeline; Medallion Architecture (Delta Lake, Synapse Analytics, Data Factory).",
         ]},
        {"name": "NIFTY 50 Market Forecasting & Financial Intelligence",
         "bullets": [
             "Multi-model equity framework (ARIMA/SARIMA/Prophet) — 30% reduction in directional error; Streamlit dashboard for 50 constituents.",
         ]},
        {"name": "SaaS Product Analytics Dashboard",
         "bullets": [
             "MRR growth $47K→$196K (318%); cohort A/B test (n=2,000) achieving 6.8pp conversion lift and +26.7% revenue lift.",
         ]},
    ]

    for proj in projects_data:
        story.append(Paragraph(f"<b>{proj.get('name','')}</b>", s_role))
        for b in (proj.get("bullets") or []):
            story.append(bullet_p(b))

    # ── Footer ─────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 5))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_LGRAY,
                             spaceAfter=2))
    score_str = f"  |  ATS keyword match: {ats_score:.0%}" if ats_score > 0 else ""
    story.append(Paragraph(
        f"<i>Application for: {job['title']} at {job['company']}"
        f"{score_str}  \u00B7  "
        f"Generated {datetime.now(timezone.utc).strftime('%d %b %Y')}</i>",
        s_footer))

    doc.build(story)


def generate_resume_for_job(j: dict) -> str:
    """
    Generate a tailored, ATS-optimised resume PDF for job j.

    Algorithm:
      1. Extract all JD keywords from title + tags + description.
      2. Call Claude to rewrite the resume embedding those keywords.
      3. Score the output — if < 95%, retry with missing keywords explicitly listed.
      4. Up to 3 attempts; uses the best-scoring version.
      5. Render final version to PDF via ReportLab.

    Returns the relative file path on success, or an error string.
    """
    if not REPORTLAB_AVAILABLE:
        return "reportlab not installed"

    Path(RESUME_OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    safe_id  = re.sub(r"[^a-z0-9_-]", "_", j["id"].lower())[:50]
    pdf_path = str(Path(RESUME_OUTPUT_DIR) / f"resume_{safe_id}.pdf")

    if Path(pdf_path).exists():
        return pdf_path

    # Build JD text from all available fields
    jd_context = (f"{j['title']} at {j['company']}. "
                  f"Location: {j['location']}. "
                  f"Tags / Required Skills: {j['tags']}.")

    # Step 1 — extract keywords
    keywords = _extract_jd_keywords(j["title"], j["tags"], jd_context)
    if not keywords:
        keywords = j["tags"].lower().split(",") if j["tags"] else []

    best_rewritten  = {}
    best_score      = 0.0
    best_missing    = keywords
    TARGET_SCORE    = 0.95
    MAX_ATTEMPTS    = 3

    if ANTHROPIC_API_KEY:
        for attempt in range(1, MAX_ATTEMPTS + 1):
            print(f"      🤖 Resume attempt {attempt}/{MAX_ATTEMPTS} — "
                  f"{j['title']} @ {j['company']}")

            missing_prev = best_missing if attempt > 1 else None
            rewritten    = _call_claude_resume(
                job_title=j["title"],
                company=j["company"],
                jd_text=jd_context,
                job_tags=j["tags"],
                keywords=keywords,
                missing_from_prev=missing_prev,
            )

            if not rewritten:
                print(f"        ⚠ Empty response on attempt {attempt}")
                time.sleep(1)
                continue

            # Step 2 — score
            resume_text   = _rewritten_to_text(rewritten)
            score, missing = _ats_score(resume_text, keywords)
            print(f"        📊 ATS score: {score:.1%}  "
                  f"({len(keywords)-len(missing)}/{len(keywords)} keywords matched)")

            if score > best_score:
                best_score     = score
                best_rewritten = rewritten
                best_missing   = missing

            if best_score >= TARGET_SCORE:
                print(f"        ✅ Target {TARGET_SCORE:.0%} reached — done")
                break

            if attempt < MAX_ATTEMPTS:
                print(f"        🔄 Below target — retrying with {len(missing)} "
                      f"missing keywords injected")
                time.sleep(1.5)

        if best_score > 0:
            print(f"      📄 Final ATS score: {best_score:.1%} "
                  f"({'✅ PASS' if best_score >= TARGET_SCORE else '⚠ BEST EFFORT'})")
        else:
            print(f"      ℹ  All API attempts failed — using base resume")
    else:
        print(f"      ℹ  No ANTHROPIC_API_KEY — base resume for: {j['title']}")

    # Step 3 — render PDF
    try:
        _build_resume_pdf(j, best_rewritten, pdf_path, ats_score=best_score)
        return pdf_path
    except Exception as e:
        print(f"      ⚠ PDF build error for {j['id']}: {e}")
        return f"PDF error: {e}"


def create_or_update_excel(new_jobs: list, filepath: str) -> None:
    if not EXCEL_AVAILABLE:
        csv_path = filepath.replace(".xlsx", ".csv")
        write_header = not Path(csv_path).exists()
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow([c[0] for c in EXCEL_COLUMNS])
            for j in new_jobs:
                cat = BOARD_CATEGORY.get(j["source"], "Other")
                resume_path = generate_resume_for_job(j)
                w.writerow([cat, j["title"], j["company"], j["location"],
                            j["source"], j["salary"], j["tags"],
                            j["posted"], j["scraped"],
                            j["url"], "", "", _match_score(j), resume_path])
        print(f"    📄 Jobs appended to {csv_path}")
        return

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
    from collections     import Counter

    thin   = Side(style="thin",   color="C8D8E8")
    medium = Side(style="medium",  color="6A9DC8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_cols = len(EXCEL_COLUMNS)
    last_col = get_column_letter(num_cols)

    # ── Load or create workbook ────────────────────────────────────────────────
    if Path(filepath).exists():
        wb = load_workbook(filepath)
        ws = wb.active
        # Find first empty row by skipping divider rows too
        start_row = ws.max_row + 1
        is_new    = False
    else:
        wb       = Workbook()
        ws       = wb.active
        ws.title = "All Jobs"
        is_new   = True

    # ── Write header only on new file ─────────────────────────────────────────
    if is_new:
        hdr_fill = PatternFill("solid", fgColor="1A2F5A")
        hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        for col_idx, (col_name, col_width) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = Border(left=medium, right=medium,
                                    top=medium,  bottom=medium)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = f"A1:{last_col}1"
        start_row = 2

    # ── Sort new jobs by category order then source ────────────────────────────
    CAT_ORDER = ["🌍 Remote Aggregator", "🏢 ATS Board", "🇮🇳 India Portal"]
    def sort_key(j):
        cat = BOARD_CATEGORY.get(j["source"], "Other")
        return (CAT_ORDER.index(cat) if cat in CAT_ORDER else 99, j["source"])
    new_jobs_sorted = sorted(new_jobs, key=sort_key)

    # ── Write rows with category group dividers ────────────────────────────────
    current_row = start_row
    prev_cat    = None

    for j in new_jobs_sorted:
        cat       = BOARD_CATEGORY.get(j["source"], "Other")
        src_color = SOURCE_COLORS.get(j["source"], "444444")
        src_fill  = PatternFill("solid", fgColor=src_color)

        # ── Category divider row when group changes ────────────────────────────
        if cat != prev_cat:
            div_bg   = CATEGORY_DIVIDER_BG.get(cat, "444444")
            div_fill = PatternFill("solid", fgColor=div_bg)
            div_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            div_border = Border(left=medium, right=medium,
                                top=medium,  bottom=medium)

            # Merge all columns for the divider label
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row,   end_column=num_cols)
            div_cell = ws.cell(row=current_row, column=1,
                               value=f"  {cat}  ──────────────────────────────────────────────────────────")
            div_cell.font      = div_font
            div_cell.fill      = div_fill
            div_cell.alignment = Alignment(horizontal="left", vertical="center")
            div_cell.border    = div_border
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            prev_cat = cat

        # ── Data row ──────────────────────────────────────────────────────────
        tints    = CATEGORY_ROW_TINT.get(cat, ("F5F5F5", "EBEBEB"))
        row_fill = PatternFill("solid", fgColor=tints[current_row % 2])

        # Generate tailored resume PDF for this job
        resume_path = generate_resume_for_job(j)

        values = [
            cat,
            j["title"], j["company"], j["location"], j["source"],
            j["salary"], j["tags"], j["posted"], j["scraped"],
            j["url"], "", "", _match_score(j),
            resume_path,   # Col 14 — Tailored Resume
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)

            # Column 1 = Board Category (italic, muted)
            if col_idx == 1:
                cell.font      = Font(name="Arial", size=8, italic=True,
                                      color=div_bg if (div_bg := CATEGORY_DIVIDER_BG.get(cat,"444444")) else "444444")
                cell.fill      = row_fill
            # Column 5 = Source badge (coloured bg, white text)
            elif col_idx == 5:
                cell.font  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                cell.fill  = src_fill
            # Match Score column (col 13)
            elif col_idx == 13:
                score = str(val)
                score_color = ("15803D" if "Excellent" in score else
                               "1D4ED8" if "Good" in score else "92400E")
                cell.font  = Font(name="Arial", size=9, bold=True, color=score_color)
                cell.fill  = row_fill
            # Tailored Resume column (col 14) — styled as download link
            elif col_idx == 14:
                cell.fill = row_fill
                if resume_path and not resume_path.startswith(("PDF error", "reportlab", "API key")):
                    cell.value     = "📄 Download Resume"
                    cell.hyperlink = resume_path
                    cell.font      = Font(name="Arial", size=9, bold=True,
                                          color="0369A1", underline="single")
                else:
                    cell.value = resume_path or "—"
                    cell.font  = Font(name="Arial", size=8, color="999999", italic=True)
            else:
                cell.font  = Font(name="Arial", size=9, color="1A2F5A")
                cell.fill  = row_fill

            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(col_idx in (2, 7, 12)),
                                       horizontal="left")
            cell.border = border

        # Apply Link (col 10) — hyperlink
        link_cell = ws.cell(row=current_row, column=10)
        if j["url"]:
            link_cell.hyperlink = j["url"]
            link_cell.value     = "▶ Apply"
            link_cell.font      = Font(name="Arial", size=9,
                                       color="1A6FCF", underline="single")

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # ── Summary sheet (refreshed each run) ────────────────────────────────────
    if "Summary" not in wb.sheetnames:
        ws2 = wb.create_sheet("Summary")
    else:
        ws2 = wb["Summary"]
    ws2.delete_rows(1, ws2.max_row + 1)

    # Title
    ws2.merge_cells("A1:C1")
    title_cell = ws2["A1"]
    title_cell.value     = "Abhishek Agasti — Job Tracker Summary"
    title_cell.font      = Font(name="Arial", bold=True, size=13, color="1A2F5A")
    title_cell.alignment = Alignment(horizontal="left")

    ws2.cell(row=2, column=1, value="Last updated:").font = Font(name="Arial", italic=True, size=10)
    ws2.cell(row=2, column=2,
             value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")).font = \
        Font(name="Arial", size=10)

    # Group + source breakdown
    write_row = 4
    hdr_fill2 = PatternFill("solid", fgColor="1A2F5A")
    for label, col_width, col_idx in [("Category / Source", 28, "A"),
                                       ("Count (this run)",  16, "B"),
                                       ("Category",          20, "C")]:
        cell = ws2.cell(row=write_row, column=ord(col_idx)-64, value=label)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill2
        ws2.column_dimensions[col_idx].width = col_width
    write_row += 1

    src_counts = Counter(j["source"] for j in new_jobs)
    cat_counts = Counter(BOARD_CATEGORY.get(j["source"],"Other") for j in new_jobs)

    for cat in CAT_ORDER:
        if cat not in cat_counts: continue
        div_bg   = CATEGORY_DIVIDER_BG.get(cat, "444444")
        div_fill = PatternFill("solid", fgColor=div_bg)
        # Category subtotal row
        for ci in range(1, 4):
            cell = ws2.cell(row=write_row, column=ci)
            cell.fill = div_fill
        ws2.cell(row=write_row, column=1,
                 value=f" {cat}").font = Font(name="Arial", bold=True,
                                              color="FFFFFF", size=10)
        ws2.cell(row=write_row, column=2,
                 value=cat_counts[cat]).font = Font(name="Arial", bold=True,
                                                    color="FFFFFF", size=10)
        write_row += 1

        # Per-source rows under category
        tints = CATEGORY_ROW_TINT.get(cat, ("F5F5F5","EBEBEB"))
        for src, cnt in sorted(src_counts.items(),
                               key=lambda x: -x[1]):
            if BOARD_CATEGORY.get(src) != cat: continue
            fill = PatternFill("solid", fgColor=tints[write_row % 2])
            for ci in range(1, 4):
                ws2.cell(row=write_row, column=ci).fill = fill
            ws2.cell(row=write_row, column=1,
                     value=f"    {src}").font = Font(name="Arial", size=10, color="1A2F5A")
            ws2.cell(row=write_row, column=2,
                     value=cnt).font = Font(name="Arial", size=10, color="1A2F5A")
            write_row += 1

    # Grand total
    total_fill = PatternFill("solid", fgColor="1A2F5A")
    for ci in range(1, 4):
        ws2.cell(row=write_row, column=ci).fill = total_fill
    ws2.cell(row=write_row, column=1,
             value="TOTAL (this run)").font = Font(name="Arial", bold=True,
                                                   color="FFFFFF", size=11)
    ws2.cell(row=write_row, column=2,
             value=len(new_jobs)).font = Font(name="Arial", bold=True,
                                              color="FFFFFF", size=11)

    wb.save(filepath)
    print(f"    📊 Excel updated: {filepath}  (+{len(new_jobs)} rows, "
          f"{len(set(BOARD_CATEGORY.get(j['source']) for j in new_jobs))} categories, "
          f"resumes in {RESUME_OUTPUT_DIR}/)")

# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL
# ═══════════════════════════════════════════════════════════════════════════════
SOURCE_META = {
    "Remotive":            {"color": "#E84040", "icon": "🌐"},
    "Jobicy":              {"color": "#F07020", "icon": "📡"},
    "We Work Remotely":    {"color": "#2563EB", "icon": "💻"},
    "Remote OK":           {"color": "#059669", "icon": "✅"},
    "Working Nomads":      {"color": "#7C3AED", "icon": "🗺"},
    "Jobspresso":          {"color": "#DB2777", "icon": "☕"},
    "AngelList/Wellfound": {"color": "#EA580C", "icon": "🚀"},
    "Surely":              {"color": "#0369A1", "icon": "🎯"},
    "Greenhouse":          {"color": "#1F6FBF", "icon": "🌱"},
    "Lever":               {"color": "#1A6B3C", "icon": "⚙️"},
    "Ashby":               {"color": "#5B3090", "icon": "🔷"},
    "Workday":             {"color": "#B84A0D", "icon": "💼"},
    "SmartRecruiters":     {"color": "#1A3566", "icon": "🎪"},
    "Instahyre":           {"color": "#0F766E", "icon": "🇮🇳"},
    "Naukri":              {"color": "#B45309", "icon": "🔍"},
    "iimjobs":             {"color": "#9D174D", "icon": "🏛"},
}
SOURCE_ORDER = list(SOURCE_META.keys())

def build_email(jobs: list, date_str: str) -> str:
    by_src: dict = {}
    for j in jobs:
        by_src.setdefault(j["source"], []).append(j)

    pills = ""
    for s in SOURCE_ORDER:
        if s not in by_src: continue
        c = SOURCE_META[s]["color"]
        i = SOURCE_META[s]["icon"]
        pills += (f'<span style="background:{c};color:#fff;border-radius:14px;'
                  f'padding:4px 12px;font-size:11px;margin:3px 4px 3px 0;'
                  f'display:inline-block;">{i} {s} <b>{len(by_src[s])}</b></span>')

    rows = ""
    for src in SOURCE_ORDER:
        grp = by_src.get(src, [])
        if not grp: continue
        c = SOURCE_META[src]["color"]
        i = SOURCE_META[src]["icon"]
        rows += (f'<tr><td colspan="4" style="background:{c};color:#fff;'
                 f'font-weight:700;padding:9px 16px;font-size:13px;">'
                 f'{i}&nbsp;&nbsp;{src} — {len(grp)} new job(s)</td></tr>')
        for j in grp:
            short = j["url"][:75] + ("…" if len(j["url"]) > 75 else "")
            score = _match_score(j)
            score_color = ("#15803D" if "Excellent" in score else
                           "#1D4ED8" if "Good" in score else "#92400E")
            rows += (
                f'<tr>'
                f'<td style="padding:8px 16px 3px;font-size:13px;font-weight:600;'
                f'color:#1A2F5A;vertical-align:top;width:38%">{j["title"]}</td>'
                f'<td style="padding:8px 10px 3px;font-size:12px;color:#555;'
                f'vertical-align:top;width:22%">{j["company"]}</td>'
                f'<td style="padding:8px 10px 3px;font-size:12px;color:#555;'
                f'vertical-align:top;width:18%">{j["location"]}</td>'
                f'<td style="padding:8px 10px 3px;font-size:11px;font-weight:700;'
                f'color:{score_color};vertical-align:top;width:14%">{score}</td>'
                f'</tr>'
                f'<tr style="border-bottom:1px solid #f0f0f0;">'
                f'<td colspan="4" style="padding:1px 16px 10px;">'
                f'<a href="{j["url"]}" style="color:{c};font-size:12px;'
                f'text-decoration:none;font-weight:500;">▶ Apply → {short}</a>'
                + (f'&nbsp;&nbsp;<span style="font-size:11px;color:#888;">'
                   f'💰 {j["salary"]}</span>' if j.get("salary") else "")
                + f'</td></tr>'
            )

    return f"""<!DOCTYPE html><html lang="en">
<body style="font-family:Arial,sans-serif;max-width:800px;margin:0 auto;
             background:#f5f7fa;color:#333;">
  <div style="background:#1A2F5A;padding:24px 28px;border-radius:8px 8px 0 0;">
    <div style="font-size:11px;color:#8AAFD8;letter-spacing:1px;
                text-transform:uppercase;margin-bottom:6px;">Daily Job Digest</div>
    <h1 style="color:#fff;margin:0;font-size:22px;font-weight:700;">
      🔔 Abhishek Agasti — Full Remote Job Tracker
    </h1>
    <p style="color:#C8DDF5;margin:7px 0 0;font-size:13px;">
      {date_str} &nbsp;·&nbsp;
      <strong style="color:#fff">{len(jobs)} new</strong> jobs across
      15 platforms · Excel tracker updated
    </p>
  </div>
  <div style="background:#E8F0FB;padding:12px 28px;">
    <div style="font-weight:600;font-size:12px;color:#1A2F5A;margin-bottom:6px;">
      New jobs by source:</div>
    {pills}
  </div>
  <div style="background:#fff;border:1px solid #D8E3F0;border-top:none;">
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;">
      <tr style="background:#EEF3FA;border-bottom:2px solid #D8E3F0;">
        <th style="padding:9px 16px;text-align:left;font-size:12px;color:#555;
                   font-weight:600;width:38%">Role</th>
        <th style="padding:9px 10px;text-align:left;font-size:12px;color:#555;
                   font-weight:600;width:22%">Company</th>
        <th style="padding:9px 10px;text-align:left;font-size:12px;color:#555;
                   font-weight:600;width:18%">Location</th>
        <th style="padding:9px 10px;text-align:left;font-size:12px;color:#555;
                   font-weight:600;width:14%">Match</th>
      </tr>
      {rows}
    </table>
  </div>
  <div style="background:#f0f2f5;padding:14px 28px;font-size:11px;color:#888;
              border-top:3px solid #E84040;border-radius:0 0 8px 8px;line-height:1.7;">
    <strong style="color:#444;">Abhishek Agasti Job Tracker v6 — Full Edition + Resume AI</strong><br>
    Aggregators: Remotive · Jobicy · We Work Remotely · Remote OK ·
    Working Nomads · Jobspresso · AngelList/Wellfound · Surely<br>
    ATS boards: Greenhouse · Lever · Ashby · Workday · SmartRecruiters (100+ boards)<br>
    India-specific: Instahyre · Naukri · iimjobs<br>
    <b>NEW v6:</b> Every job now has an ATS-optimised tailored resume PDF (Col N in Excel)
    — Claude rewrites your base resume for 90%+ keyword match with each JD.<br>
    Download <em>jobs_tracker.xlsx</em> + <em>resumes/</em> folder from the Actions artifact.
  </div>
</body></html>"""


def send_email(subject: str, html: str) -> None:
    if not EMAIL_SENDER or not EMAIL_PASSWORD:
        out = Path("digest_output.html")
        out.write_text(html, encoding="utf-8")
        print(f"\n  ⚠  No credentials — digest saved to {out}")
        return
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = EMAIL_RECIPIENT
    msg.attach(MIMEText(html, "html"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(EMAIL_SENDER, EMAIL_PASSWORD)
        s.sendmail(EMAIL_SENDER, EMAIL_RECIPIENT, msg.as_string())
    print(f"\n  ✅  Email sent → {EMAIL_RECIPIENT}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    W = 72
    print("=" * W)
    print(f"  Abhishek Job Scraper v7 — FULL EDITION (16 sources + Resume AI 95%+)")
    print(f"  {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * W)
    seen = load_seen()
    print(f"\n  Previously seen: {len(seen)} jobs\n")

    all_jobs: list = []

    print("[ 1/16  Remotive ]")
    all_jobs.extend(scrape_remotive())

    print("\n[ 2/16  Jobicy ]")
    all_jobs.extend(scrape_jobicy())

    print("\n[ 3/16  We Work Remotely ]")
    all_jobs.extend(scrape_weworkremotely())

    print("\n[ 4/16  Remote OK ]")
    all_jobs.extend(scrape_remoteok())

    print("\n[ 5/16  Working Nomads ]")
    all_jobs.extend(scrape_workingnomads())

    print("\n[ 6/16  Jobspresso ]")
    all_jobs.extend(scrape_jobspresso())

    print("\n[ 7/16  AngelList / Wellfound ]")
    all_jobs.extend(scrape_wellfound())

    print("\n[ 8/16  Surely ]")
    all_jobs.extend(scrape_surely())

    print(f"\n[ 9/16  Greenhouse — {len(GREENHOUSE_SLUGS)} boards ]")
    all_jobs.extend(scrape_greenhouse_all())

    print(f"\n[ 10/16 Lever — {len(LEVER_SLUGS)} boards ]")
    all_jobs.extend(scrape_lever_all())

    print(f"\n[ 11/16 Ashby — {len(ASHBY_SLUGS)} boards ]")
    all_jobs.extend(scrape_ashby_all())

    print(f"\n[ 12/16 Workday — {len(WORKDAY_BOARDS)} tenants ]")
    all_jobs.extend(scrape_workday_all())

    print(f"\n[ 13/16 SmartRecruiters — {len(SMARTRECRUITERS_SLUGS)} companies ]")
    all_jobs.extend(scrape_smartrecruiters_all())

    print("\n[ 14/16 Instahyre (India) ]")
    all_jobs.extend(scrape_instahyre())

    print("\n[ 15/16 Naukri (India) ]")
    all_jobs.extend(scrape_naukri())

    print("\n[ 16/16 iimjobs (India premium) ]")
    all_jobs.extend(scrape_iimjobs())

    # ── Deduplicate ───────────────────────────────────────────────────────────
    print(f"\n  Total matching this run   : {len(all_jobs)}")
    this_run: set = set()
    new_jobs = []
    for j in all_jobs:
        if j["id"] not in seen and j["id"] not in this_run:
            new_jobs.append(j); this_run.add(j["id"])
    print(f"  New (not seen before)     : {len(new_jobs)}")
    seen.update(this_run)
    save_seen(seen)

    if not new_jobs:
        print("\n  No new jobs today — no email sent.\n"); return

    # ── Excel ─────────────────────────────────────────────────────────────────
    print(f"\n  Updating Excel tracker …")
    create_or_update_excel(new_jobs, EXCEL_FILE)

    # ── Email ─────────────────────────────────────────────────────────────────
    date_str = datetime.now(timezone.utc).strftime("%d %b %Y")
    subject  = (f"🔔 {len(new_jobs)} New Job Match(es) — {date_str} | "
                f"Abhishek Job Tracker v5")
    send_email(subject, build_email(new_jobs, date_str))

    print("\n  ── Breakdown by source ────────────────────────────")
    from collections import Counter
    counts = Counter(j["source"] for j in new_jobs)
    for s in SOURCE_ORDER:
        if s in counts:
            print(f"    {s:<28} {counts[s]}")
    print("=" * W)


if __name__ == "__main__":
    main()
